from langchain.chat_models import ChatOpenAI
from langchain.prompts.chat import ChatPromptTemplate, SystemMessagePromptTemplate, HumanMessagePromptTemplate
from langchain.chains import LLMChain
from pydantic import BaseModel, Field
from typing import List, Optional, Union, Any, Literal
from openai import OpenAI
import pandas as pd
from collections import defaultdict
import networkx as nx
import matplotlib.pyplot as plt
from collections import defaultdict
from uuid import uuid4
import xml.etree.ElementTree as ET
import math
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from docx import Document
from glob import glob



# Initialize the language model
llm = ChatOpenAI(
    model_name="gpt-4o",
    temperature=0,
    openai_api_key="" ##ADD KEY
)
client = OpenAI(api_key="") ## add key




def add_dummy(flowchart: List[dict]) -> List[dict]:
    for el in flowchart:
        if "domanda" in el:
            for ramo in el["rami"]:
                if ramo["esito"] == "FINE" and not ramo["ramo"]:
                    ramo["ramo"].append({"soggetto": "Sistema", "azioni": ["Fine"]})
                if ramo["esito"] == "CONTINUA" and not ramo["ramo"]:
                    ramo["ramo"].append({"soggetto": "Sistema", "azioni": ["Continua"]})
                else:
                    add_dummy(ramo["ramo"])
    return flowchart

def add_ids(flowchart: List[dict]) -> List[dict]:
    activity_counter = 1
    condition_counter = 1
    ramo_counter = 1

    def _add_ids(element: Any):
        nonlocal activity_counter, condition_counter, ramo_counter

        if "soggetto" in element:
            element["ID"] = [f"A{activity_counter + i}" for i in range(len(element["azioni"]))]
            activity_counter += len(element["azioni"])

        elif "domanda" in element:
            element["ID"] = f"C{condition_counter}"
            condition_counter += 1

            for ramo in element["rami"]:
                ramo["ID"] = f"R{ramo_counter}"
                ramo_counter += 1
                for sub_element in ramo["ramo"]:
                    _add_ids(sub_element)

    for el in flowchart:
        _add_ids(el)

    return flowchart


def add_successor_ids(flowchart: List[dict]) -> List[dict]:
    id_to_element = {}

    def index_elements(elements: List[Any]):
        for el in elements:
            if "soggetto" in el:
                for i, id_azione in enumerate(el["ID"]):
                    id_to_element[id_azione] = {"parent": el, "index": i, "type": "azione"}
            elif "domanda" in el:
                id_to_element[el["ID"]] = {"parent": el, "type": "condizione"}
                for ramo in el["rami"]:
                    id_to_element[ramo["ID"]] = {"parent": ramo, "type": "ramo"}
                    index_elements(ramo["ramo"])

    index_elements(flowchart)

    def find_next_id(context, idx):
        for i in range(idx + 1, len(context)):
            el = context[i]
            if "soggetto" in el:
                return el["ID"][0]
            elif "domanda" in el:
                return el["ID"]
        return None

    def process(elements: List[Any], next_chain=None, context_stack=None):
        if context_stack is None:
            context_stack = []

        context_stack.append(elements)

        for idx, el in enumerate(elements):
            if "soggetto" in el:
                for i, curr_id in enumerate(el["ID"]):
                    if i < len(el["ID"]) - 1:
                        succ = [el["ID"][i + 1]]
                    else:
                        next_id = find_next_id(elements, idx) or next_chain
                        succ = [next_id] if next_id else []
                    id_to_element[curr_id]["parent"].setdefault("ID SUCCESSIVI", {})[curr_id] = succ

            elif "domanda" in el:
                el["ID SUCCESSIVI"] = [ramo["ID"] for ramo in el["rami"]]

                for ramo in el["rami"]:
                    next_id = find_next_id(elements, idx) or next_chain if ramo["esito"] in ["CONTINUA", "EREDITATO"] else None

                    if ramo["ramo"]:
                        primo = ramo["ramo"][0]
                        ramo["ID SUCCESSIVI"] = [primo["ID"][0]] if "soggetto" in primo else [primo["ID"]]

                        process(ramo["ramo"], next_chain=next_id, context_stack=context_stack)

                        ultimo = ramo["ramo"][-1]
                        if ramo["esito"] in ["CONTINUA", "EREDITATO"]:
                            if "soggetto" in ultimo:
                                ultimo_id = ultimo["ID"][-1]
                                succ = [next_id] if next_id else []
                                id_to_element[ultimo_id]["parent"].setdefault("ID SUCCESSIVI", {})[ultimo_id] = succ
                            elif "domanda" in ultimo:
                                ultimo_id = ultimo["ID"]
                                id_to_element[ultimo_id]["parent"].setdefault("ID SUCCESSIVI", [])
                                if next_id and next_id not in id_to_element[ultimo_id]["parent"]["ID SUCCESSIVI"]:
                                    id_to_element[ultimo_id]["parent"]["ID SUCCESSIVI"].append(next_id)
                    else:
                        ramo["ID SUCCESSIVI"] = [next_id] if next_id else []

        context_stack.pop()

    process(flowchart)
    return flowchart
def aggiungi_colonna_precedenti(df: pd.DataFrame) -> pd.DataFrame:
    precedenti = {}

    for _, row in df.iterrows():
        id_corrente = row["ID"]
        successivi = row["ID SUCCESSIVI"]
        if isinstance(successivi, list):
            for succ in successivi:
                if succ not in precedenti:
                    precedenti[succ] = []
                precedenti[succ].append(id_corrente)

    df["ID PRECEDENTI"] = df["ID"].apply(lambda x: precedenti.get(x, []))
    return df

def json_to_dataframe(flowchart: List[dict]) -> pd.DataFrame:
    rows = []

    def extract(elements: List[Any]):
        for el in elements:
            if "soggetto" in el:
                for i, id_ in enumerate(el["ID"]):
                    rows.append({
                        "NOME": el["azioni"][i],
                        "SOGGETTO": el["soggetto"],
                        "ID": id_,
                        "ID SUCCESSIVI": el.get("ID SUCCESSIVI", {}).get(id_, [])
                    })
            elif "domanda" in el:
                successori = [sid for sid in el.get("ID SUCCESSIVI", []) if isinstance(sid, str) and sid.startswith("R")]

                rows.append({
                    "NOME": el["domanda"],
                    "SOGGETTO": None, 
                    "ID": el["ID"],
                    "ID SUCCESSIVI": successori
                })
                for ramo in el["rami"]:
                    rows.append({
                        "NOME": ramo["risposta"],
                        "SOGGETTO": None,
                        "ID": ramo["ID"],
                        "ID SUCCESSIVI": ramo.get("ID SUCCESSIVI", [])
                    })
                    extract(ramo["ramo"])

    extract(flowchart)
    return aggiungi_colonna_precedenti(pd.DataFrame(rows))


def insert_start_events(process, sequence_flows, elements_info, coords, offset_y=+1000):
    start_events = []
    new_sequence_flows = sequence_flows.copy()

    all_targets = set(t for _, t, _ in sequence_flows)
    all_sources = set(s for s, _, _ in sequence_flows)
    
    elementi_senza_predecessori = all_sources - all_targets

    for idx, target in enumerate(sorted(elementi_senza_predecessori)):
        start_id = f"StartEvent_{uuid4().hex[:8]}"
        start_event = ET.Element("bpmn:startEvent", {"id": start_id, "name": f"Start {idx+1}"})
        process.append(start_event)
        start_events.append((start_id, "startEvent"))

        new_sequence_flows.append((start_id, target, None))

        x = 0
        y = idx * offset_y  
        coords[start_id] = (x, y)

    elements_info.extend(start_events)

    return new_sequence_flows, elements_info, coords

def insert_join_gateways(process, sequence_flows):
    from collections import defaultdict
    predecessors_map = defaultdict(list)
    for source, target, label in sequence_flows:
        predecessors_map[target].append((source, label))

    new_sequence_flows = []
    join_gateways = []

    for target, sources in predecessors_map.items():
        if len(sources) <= 1:
            new_sequence_flows.extend([(s, target, l) for s, l in sources])
            continue

        join_id = f"J_{uuid4().hex[:8]}"
        join_gateway = ET.Element("bpmn:exclusiveGateway", {"id": join_id }) 
        process.append(join_gateway)
        join_gateways.append((join_id, "exclusiveGateway"))

        for source, label in sources:
            new_sequence_flows.append((source, join_id, label))

        new_sequence_flows.append((join_id, target, None))

    return new_sequence_flows, join_gateways

def create_bpmn_from_df(df: pd.DataFrame, filename: str):
    def bpmn_id(prefix): 
        return f"{prefix}_{uuid4().hex[:8]}"
    
    bpmn_ns = {
        "bpmn": "http://www.omg.org/spec/BPMN/20100524/MODEL",
        "xsi": "http://www.w3.org/2001/XMLSchema-instance"
    }

    ET.register_namespace("bpmn", bpmn_ns["bpmn"])
    ET.register_namespace("xsi", bpmn_ns["xsi"])

    definitions = ET.Element("bpmn:definitions", {
        "id": "Definitions_1",
        "targetNamespace": "http://example.bpmn",
        "xmlns:bpmn": bpmn_ns["bpmn"],
        "xmlns:xsi": bpmn_ns["xsi"],
        "xmlns:bpmndi": "http://www.omg.org/spec/BPMN/20100524/DI",
        "xmlns:omgdc": "http://www.omg.org/spec/DD/20100524/DC",
        "xmlns:omgdi": "http://www.omg.org/spec/DD/20100524/DI"
    })

    process = ET.SubElement(definitions, "bpmn:process", {
        "id": "Process_1",
        "isExecutable": "true"
    })

    end_event_ids = [bpmn_id("EndEvent") for _ in range(10)]


    task_elements = []
    gateway_elements = []
    annotation_elements = []
    association_elements = []
    end_events = []
    sequence_flows = []
    task_map = {}  

    for _, row in df.iterrows():
        id_ = str(row["ID"])
        name = row["NOME"] if pd.notna(row["NOME"]) else ""

        if id_.startswith("A"):
            el = ET.Element("bpmn:task", {"id": id_, "name": name})
            task_elements.append(el)
            task_map[id_] = id_

        elif id_.startswith("C"):
            el = ET.Element("bpmn:exclusiveGateway", {"id": id_, "name": name})
            gateway_elements.append(el)
            task_map[id_] = id_

    for _, row in df.iterrows():
        id_ = str(row["ID"])
        successors = row["ID SUCCESSIVI"] if isinstance(row["ID SUCCESSIVI"], list) else []
        predecessors = row["ID PRECEDENTI"] if isinstance(row["ID PRECEDENTI"], list) else []

        if id_.startswith("R"):
            continue


        for succ in successors:
            if succ.startswith("R"):
                ramo_row = df[df["ID"] == succ].squeeze()
                ramo_name = ramo_row["NOME"] if pd.notna(ramo_row["NOME"]) else ""
                ramo_successori = ramo_row["ID SUCCESSIVI"] if isinstance(ramo_row["ID SUCCESSIVI"], list) else []
                for real_target in ramo_successori:
                    sequence_flows.append((id_, real_target, ramo_name))
            else:
                sequence_flows.append((id_, succ, None))

        if not successors and end_event_ids:
            this_end_id = end_event_ids.pop(0)
            end_event = ET.Element("bpmn:endEvent", {"id": this_end_id, "name": "End"})
            end_events.append(end_event)
            sequence_flows.append((id_, this_end_id, None))

    for el in task_elements + gateway_elements + end_events + annotation_elements + association_elements:
        process.append(el)
    sequence_flows, join_gateway_info = insert_join_gateways(process, sequence_flows)

    for source, target, label in sequence_flows:
        sf_id = f"{source}_{target}"
        sf_attribs = {
            "id": sf_id,
            "sourceRef": source,
            "targetRef": target
        }
        if label:
            sf_attribs["name"] = label
        ET.SubElement(process, "bpmn:sequenceFlow", sf_attribs)

    elements_info =   [(t.attrib["id"], "task") for t in task_elements] +         [(g.attrib["id"], "exclusiveGateway") for g in gateway_elements] +         [(e.attrib["id"], "endEvent") for e in end_events] +         [(f"Annotation_{row['ID']}", "textAnnotation")
         for _, row in df.iterrows()
         if pd.notna(row.get("SOGGETTO")) and str(row["SOGGETTO"]).strip()] + \
        join_gateway_info

    coords = {}
    sequence_flows, elements_info, coords = insert_start_events(process, sequence_flows, elements_info, coords) 


    add_bpmn_diagram_elements(definitions, elements_info, [(s, t) for s, t, _ in sequence_flows], coords)

    tree = ET.ElementTree(definitions)
    ET.indent(tree, space="  ", level=0)
    tree.write(filename, encoding="utf-8", xml_declaration=True)

    with open(filename, "r", encoding="utf-8") as f:
        bpmn_content = f.read()


    task_info_dict = {}

    for _, row in df.iterrows():
        id_ = str(row["ID"])

        if id_.startswith("A"):
            def sval(key):
                v = row.get(key, "")
                if pd.isna(v):
                    return ""
                v = str(v).strip()
                if v.lower() in {"n/a", "na", "nan"}:
                    return ""
                return v
            info_parts = [f"ID: {id_}"]
            description       = sval("descrizione")
            actor             = sval("attore")
            recipient         = sval("destinatario")
            legal_reference   = sval("riferimento_normativo")
            output            = sval("output")
            documents         = sval("documenti")
            time_reference    = sval("riferimento_temporale")
            forms             = sval("moduli")
            
            if description:     info_parts.append(f"DESCRIZIONE: {description}")
            if actor:           info_parts.append(f"ATTORE: {actor}")
            if recipient:       info_parts.append(f"RICEVENTE: {recipient}")
            if legal_reference: info_parts.append(f"RIFERIMENTO NORMATIVO: {legal_reference}")
            if output:          info_parts.append(f"OUTPUT: {output}")
            if documents:       info_parts.append(f"DOCUMENTI: {documents}")
            if time_reference:  info_parts.append(f"RIFERIMENTO TEMPORALE: {time_reference}")
            if forms:           info_parts.append(f"MODULI: {forms}")

            task_info_dict[id_] = "<br>".join(info_parts)

    html_template = html_template = '''
<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <title>BPMN Viewer - Informazioni a Comparsa</title>
  <script src="https://unpkg.com/bpmn-js@11.5.0/dist/bpmn-viewer.development.js"></script>
  <style>
    html, body {{
      height: 100%;
      margin: 0;
      overflow: hidden;
    }}
    #canvas {{
      width: 100%;
      height: 100%;
      border: 1px solid #ccc;
      position: relative;
    }}
    #infoBox {{
      position: absolute;
      background: #fff;
      border: 1px solid #888;
      padding: 10px;
      box-shadow: 2px 2px 8px rgba(0,0,0,0.3);
      display: none;
      z-index: 10;
      max-width: 300px;
      font-family: Arial, sans-serif;
      font-size: 14px;
    }}
  </style>
</head>
<body>
<div id="canvas"></div>
<div id="infoBox"></div>
<script>
const bpmnXML = {bpmn_content_json};
const viewer = new BpmnJS({{ 
  container: '#canvas'
}});
const taskInfo = {task_info_json};

viewer.importXML(bpmnXML).then(() => {{
  viewer.get('canvas').zoom('fit-viewport');
  
  const eventBus = viewer.get('eventBus');
  const canvasContainer = document.getElementById('canvas');
  const infoBox = document.getElementById('infoBox');
  const canvas = viewer.get('canvas');
  
  let isDragging = false;
  let lastX = 0;
  let lastY = 0;
  
  canvasContainer.addEventListener('mousedown', function(event) {{
    if (event.button === 0) {{
      isDragging = true;
      lastX = event.clientX;
      lastY = event.clientY;
      canvasContainer.style.cursor = 'grabbing';
      event.preventDefault();
    }}
  }});
  
  canvasContainer.addEventListener('mousemove', function(event) {{
    if (isDragging) {{
      const deltaX = event.clientX - lastX;
      const deltaY = event.clientY - lastY;
      
      const viewbox = canvas.viewbox();
      canvas.viewbox({{
        x: viewbox.x - deltaX,
        y: viewbox.y - deltaY,
        width: viewbox.width,
        height: viewbox.height
      }});
      
      lastX = event.clientX;
      lastY = event.clientY;
    }}
  }});
  
  canvasContainer.addEventListener('mouseup', function() {{
    isDragging = false;
    canvasContainer.style.cursor = 'grab';
  }});
  
  canvasContainer.addEventListener('mouseleave', function() {{
    isDragging = false;
    canvasContainer.style.cursor = 'grab';
  }});
  
  canvasContainer.style.cursor = 'grab';
  
  eventBus.on('element.dblclick', function(event) {{
    const element = event.element;
    if (element.type === 'bpmn:Task') {{
      const info = taskInfo[element.id] || "Nessuna informazione disponibile.";
      const rect = canvasContainer.getBoundingClientRect();
      const x = event.originalEvent.clientX - rect.left;
      const y = event.originalEvent.clientY - rect.top;
      infoBox.style.left = x + 'px';
      infoBox.style.top = y + 'px';
      infoBox.innerHTML = `<strong>${{element.businessObject.name}}</strong><br>${{info}}`;
      infoBox.style.display = 'block';
    }}
  }});
  
  canvasContainer.addEventListener('click', () => {{
    infoBox.style.display = 'none';
  }});
  
}}).catch(err => {{
  console.error('Errore nel caricamento del diagramma', err);
}});
</script>
</body>
</html>
'''

    bpmn_content_json = json.dumps(bpmn_content, ensure_ascii=False)
    task_info_json = json.dumps(task_info_dict, ensure_ascii=False)

    html_content = html_template.format(
        bpmn_content_json=bpmn_content_json,
        task_info_json=task_info_json
    )

    return bpmn_content, html_content




def add_bpmn_diagram_elements(definitions, elements_info, sequence_flows, coords):
    bpmndi_ns = {
        "bpmndi": "http://www.omg.org/spec/BPMN/20100524/DI",
        "omgdc": "http://www.omg.org/spec/DD/20100524/DC",
        "omgdi": "http://www.omg.org/spec/DD/20100524/DI"
    }

    for prefix, uri in bpmndi_ns.items():
        ET.register_namespace(prefix, uri)

    diagram = ET.SubElement(definitions, "bpmndi:BPMNDiagram", {"id": "BPMNDiagram_1"})
    plane = ET.SubElement(diagram, "bpmndi:BPMNPlane", {
        "id": "BPMNPlane_1",
        "bpmnElement": "Process_1"
    })

    shape_size = (120, 120) 


    predecessors_map = {}
    for source, target in sequence_flows:
        predecessors_map.setdefault(target, []).append(source)

    def get_element_type(element_id):
        for eid, etype in elements_info:
            if eid == element_id:
                return etype
        return "task"  # fallback

    split_positions = {}
    def compute_position(el_id):
        if el_id in coords:
            return coords[el_id]

        el_type = get_element_type(el_id)

        if el_type == "textAnnotation":
            task_id = el_id.replace("Annotation_", "")
            if task_id not in coords:
                compute_position(task_id)  # Assicurati che il task sia posizionato prima
            task_x, task_y = coords[task_id]
            x = task_x
            y = task_y + 100
            coords[el_id] = (x, y)
            return x, y

        preds = predecessors_map.get(el_id, [])
        if not preds:
            x, y = 100, 100

        elif len(preds) == 1:
            pred = preds[0]
            px, py = compute_position(pred)
            pred_type = get_element_type(pred)

            is_split_gateway = pred_type in ("exclusiveGateway", "inclusiveGateway", "parallelGateway") and                                sum(1 for s, _ in sequence_flows if s == pred) > 1

            if is_split_gateway:
                x = px + 200
                if pred not in split_positions:
                    n_rami = sum(1 for s, t in sequence_flows if s == pred)
                    split_positions[pred] = {
                        "base_y": py - 150 * (n_rami // 2),
                        "index": 0
                    }
                offset_data = split_positions[pred]
                temp_y = offset_data["base_y"] + offset_data["index"] * 150

                while temp_y == py:  
                    offset_data["index"] += 1
                    temp_y = offset_data["base_y"] + offset_data["index"] * 150

                y = temp_y
                offset_data["index"] += 1

            else:
                x = px + 200
                y = py
        elif len(preds) > 1:
            pred_positions = [compute_position(p) for p in preds]
            
            max_x = max(pos[0] for pos in pred_positions)
            avg_y = math.floor(sum(pos[1] for pos in pred_positions) / len(pred_positions) / 10) * 10 

            x = max_x + 200  
            y = avg_y  


            for pred in preds:
                pred_type = get_element_type(pred)
                is_split_gateway = pred_type in ("exclusiveGateway", "inclusiveGateway", "parallelGateway") and                                    sum(1 for s, _ in sequence_flows if s == pred) > 1
                if is_split_gateway:
                    px, py = coords[pred]
                    x = px + 600
                    y = py

        coords[el_id] = (x, y)
        return x, y

    for el_id, el_type in elements_info:
        shape = ET.SubElement(plane, "bpmndi:BPMNShape", {
            "id": f"{el_id}_di",
            "bpmnElement": el_id
        })

        if el_type == "textAnnotation":
            width, height =  160, 35
        else:
            width, height = shape_size
        x, y = compute_position(el_id)

        ET.SubElement(shape, "omgdc:Bounds", {
            "x": str(x),
            "y": str(y),
            "width": str(width),
            "height": str(height)
        })


    if el_type in ("exclusiveGateway", "inclusiveGateway", "parallelGateway"):
        label = ET.SubElement(shape, "bpmndi:BPMNLabel")
        ET.SubElement(label, "omgdc:Bounds", {
            "x": str(x + width / 2 - 30),  
            "y": str(y - 20),              
            "width": "60",
            "height": "20"
        })

    for source, target in sequence_flows:
        sf_id = f"{source}_{target}"

        edge = ET.SubElement(plane, "bpmndi:BPMNEdge", {
            "id": f"{sf_id}_di",
            "bpmnElement": sf_id
        })

        sx, sy = coords[source]
        tx, ty = coords[target]

        sw, sh = shape_size
        tw, th = shape_size

        ET.SubElement(edge, "omgdi:waypoint", {
            "x": str(sx + sw),
            "y": str(sy + sh / 2)
        })
        ET.SubElement(edge, "omgdi:waypoint", {
            "x": str(tx),
            "y": str(ty + th / 2)
        })



    return plane, coords



def adjust_nome(row, id_to_nome):
    id_val = str(row["ID"])
    if id_val.startswith("R"):
        id_precedenti = row["ID PRECEDENTI"]
        if pd.notna(id_precedenti):
            cleaned = str(id_precedenti).strip("[]").replace("'", "")
            first_prec = cleaned.split(",")[0].strip()
            if first_prec in id_to_nome:
                cond_nome = id_to_nome[first_prec]
                return f"{cond_nome} - {row['NOME']}"
    return row["NOME"] 
def salva_documenti(csv_df, bpmn_content, html_content, nome, path):
    folder_path = os.path.join(path, nome)
    os.makedirs(folder_path, exist_ok=True)

    csv_path = os.path.join(folder_path, f"{nome}.csv")
    excel_path = os.path.join(folder_path, f"{nome}.xlsx")
    
    csv_df = csv_df.copy()
    csv_df.columns = [col.upper().replace('_', ' ') for col in csv_df.columns]

    if "ID" in csv_df.columns and "NOME" in csv_df.columns and "ID PRECEDENTI" in csv_df.columns:
        id_to_nome = dict(zip(csv_df["ID"], csv_df["NOME"]))

        csv_df["NOME"] = csv_df.apply(lambda row: adjust_nome(row, id_to_nome), axis=1)

        csv_df["NOME"] = csv_df.apply(
            lambda row: row["NOME"] if str(row["ID"]).startswith(("C", "R")) else None,
            axis=1
        )
    csv_df = csv_df.rename(columns={"NOME": "CONDIZIONE"})
    desired_order = [ "ATTIVITÀ", "CONDIZIONE", "ID", "ID SUCCESSIVI", "ID PRECEDENTI", "RIFERIMENTO NORMATIVO", "ATTORE", "DESTINATARIO", "DESCRIZIONE"]

    ordered_cols = [col for col in desired_order if col in csv_df.columns]
    remaining_cols = [col for col in csv_df.columns if col not in ordered_cols]
    csv_df = csv_df[ordered_cols + remaining_cols]
    
    csv_df.to_excel(excel_path, index=False, engine='openpyxl')

    wb = load_workbook(excel_path)
    ws = wb.active

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30     
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 30

    wb.save(excel_path)

    bpmn_path = os.path.join(folder_path, f"{nome}.bpmn")
    with open(bpmn_path, "w", encoding="utf-8") as f:
        f.write(bpmn_content)

    html_path = os.path.join(folder_path, f"{nome}.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)



class Task(BaseModel): 
    soggetto: str = Field(..., description="Chi compie l'azione")
    azioni: List[str] = Field(..., description="Lista di azioni sequenziali svolte dallo stesso soggetto.")


class Gateway(BaseModel):
    domanda: str = Field(..., description="Domanda che determina le ramificazioni del processo")
    rami: List["ConditionalFlow"] = Field(..., description="Elenco dei possibili rami in base alla risposta")

class ConditionalFlow(BaseModel):
    risposta: str = Field(..., description="Risposta alla condizione")
    ramo: List[Union[Task, Gateway]] = Field(..., description="Task o ulteriori Gateway da eseguire")
    esito: Literal["CONTINUA", "FINE", "EREDITATO"] = Field(
        ..., 
        description="Esito del ConditionalFlow: 'CONTINUA' se torna al flusso esterno; 'FINE' se porta a un endEvent; 'EREDITATO' se dipende da gateway annidati."
    )

class Flowchart(BaseModel):
    passaggi: List[Union[Task, Gateway]] = Field(..., description="Lista ordinata delle azioni e condizioni")

Flowchart.model_rebuild()

system_msg_1 = """#SYSTEM
**Contesto**: Sei un esperto nell’analisi e modellazione di procedure amministrative.

**Obiettivo**: Leggi attentamente un articolo di legge e rappresenta in modo chiaro, rigoroso e ordinato i passaggi che compongono la procedura amministrativa, distinguendo tra Task, Azioni, Gateways e ConditionalFlow, secondo una struttura facilmente leggibile e pronta per essere trasformata in un diagramma grafico (es. BPMN o flowchart).

**Input**: <ARTICOLO DI LEGGE>, il testo normativo che descrive la procedura amministrativa.

**Output**: <FLOWCHART TESTUALE>, strutturato in formato testuale.

**Approccio STEP-BY-STEP**:
    Segui rigorosamente questi passaggi:
    1. Leggi attentamente l’<ARTICOLO DI LEGGE> nella sua interezza per comprenderne il contesto generale e lo scopo della procedura;
    2. Individua le fasi che descrivono la procedura amministrativa;
    3. Comprendi la sequenza logica degli elementi della procedura;
    4. Modellazione delle fasi: Costruisci uno schema logico delle Fasi e Sottofasi della procedura;
    5. Modellazione di dettaglio, riporta il modello senza trascurare nessun passaggio:
        a. Task: insieme di azioni sequenziali svolte dallo stesso soggetto
            a.1 Azione: azione svolta da un soggetto (da esprimere con poche parole).
        b. Gateways: individua le domande o situazioni che determinano ramificazioni nella procedura.
        c. ConditionalFlow: le altenative che portano alle diramazioni dei Gateways (es. Sì/No, Facoltativo, Se dichiarato interesse, ecc.) e che si sviluppano nelle loro fasi interne con elementi della procedura (Task, Azioni, Gateways).
    6. Controllo finale:
        a. Verifica completezza, coerenza e conformità del <FLOWCHART TESTUALE> con l'<ARTICOLO DI LEGGE>;

**Vincoli e Stile**:
    - Solo informazioni essenziali alla sequenza logica: Il modello deve contenere esclusivamente le informazioni utili a ricostruire la sequenza logica della procedura. Non includere elementi accessori come tempistiche specifiche o dettagli che non sono essenziali alla comprensione della logica della procedura.
    - Indentazione per Gateways e ConditionalFlow: Utilizza le indentazioni per rappresentare Gateways e i rispettivi ConditionalFlow 
    - Possibilità di Gateways annidati: Ci possono essere Gateways annidati, quindi percorso procedurali annidati. 
    - Segnala gli EndEvent: Segnala quando una ramificazione conduce alla fine/interruzione della procedura
    - Mantieni descrizioni brevi e precise
    - Il modello deve essere concettualmente un diagramma grafico
    - Proporzionalità: L'estensione del modello deve essere proporzionata all'estensione e alla complessità dell'<ARTICOLO DI LEGGE> ;

**Audience**: Il destinatario è un analista tecnico di procedure amministrative, quindi il <FLOWCHART TESTUALE> deve essere perfetto.
"""

human_msg_1 = """#INPUT
<ARTICOLO DI LEGGE>:
{law_article}

#OUTPUT
<FLOWCHART>:"""

prompt_1 = ChatPromptTemplate.from_messages([
    SystemMessagePromptTemplate.from_template(system_msg_1),
    HumanMessagePromptTemplate.from_template(human_msg_1)
])

chain1 = LLMChain(llm=llm, prompt=prompt_1, output_key="flowchart")


system_msg_2 = """#SYSTEM
**Contesto**: Sei un esperto nell’analisi e controllo di procedure amministrative.

**Obiettivo**: Leggi attentamente un <ARTICOLO DI LEGGE> e verifica la correttezza di un flowchart che descrive la relativa procedura amministrativa. Se necessario, correggi gli errori e fornisci il flowchart definitivo, rappresentato in formato JSON preciso e conforme.

**Input**:
    - <ARTICOLO DI LEGGE>: il testo normativo che descrive la procedura amministrativa.
    - <FLOWCHART>: il flowchart testuale della procedura amministrativa, estratto dall'<ARTICOLO DI LEGGE>.

**Output**: 
    <FLOWCHART DEFINITIVO>: il flowchart corretto e definitivo, rappresentato in formato JSON conforme alla seguente struttura concettuale:

        - Task:
            - "soggetto" (stringa): Chi compie l'azione (es. "Ente pubblico").
            - "azioni" (lista di stringhe): Le azioni svolte in sequenza.

        - Gateway:
            - "domanda" (stringa): La domanda o il punto decisionale.
            - "rami" (lista di ConditionalFlow): Le alternative in base alla risposta.

        - ConditionalFlow:
            - "risposta" (stringa): La risposta o condizione.
            - "ramo" (lista ricorsiva): Gli elementi successivi (Task o ulteriori Condizioni).
            - "esito" (Literal[CONTINUA, FINE, EREDITATO]): Indica l’esito del ConditionalFlow: 'CONTINUA' se torna al flusso esterno; 'FINE' se porta a un endEvent; 'EREDITATO' se dipende da gateway annidati.

        - Flowchart:
            - Lista ordinata e strutturata degli elementi della procedura.

**Approccio STEP-BY-STEP**:
    Segui rigorosamente questi passaggi:
    1. Leggi attentamente l’<ARTICOLO DI LEGGE> per comprenderne il contesto e la sequenza logica della procedura;
    2. Individua le Task, Gateway e ConditionalFlow che descrivono la procedura amministrativa;
    3. Analizza il <FLOWCHART> proposto;
    4. Confronta in modo preciso e rigoroso il <FLOWCHART> con la procedura dall'<ARTICOLO DI LEGGE> che hai estratto;
    5. Per ogni elemento del <FLOWCHART>:
        a. Verifica se è:
            - Corretto: lo lasci invariato;
            - Errato, impreciso o incompleto: lo correggi;
            - Mancante: lo aggiungi;
            - Superfluo: lo elimini;
    6. Controllo finale:
        a. Verifica completezza, coerenza e conformità al formato richiesto;
        b. Assicurati che il <FLOWCHART FINALE> descriva perfettamente la procedura amministrativa;
        c. Non introdurre interpretazioni personali o elementi non esplicitamente presenti nell'<ARTICOLO DI LEGGE>.

**Vincoli e Stile**:
    - Formula azioni: Descrivi le azioni nella forma "nome deverbale (azione) + complemento di specificazione"
    - Descrizioni sintetiche: Utilizza descrizioni brevi ma chiare e precis
    - Formato preciso: Il <FLOWCHART FINALE> deve essere conforme alla struttura specificata

**Audience**: Il destinatario è un analista tecnico di procedure amministrative, quindi il <FLOWCHART FINALE> deve essere preciso, corretto e pronto per la modellazione grafica.

#ESEMPIO DI ELEMENTI DEL <FLOWCHART STRUTTURATO>:
    Di seguito ti do un po' di esempi di elementi di flowchart strutturato in formato JSON, conformi allo schema previsto.  
    Gli esempi servono solo per mostrarti alcune strutture tecniche e la forma corretta degli elementi, non il contenuto specifico (per il contenuto del <FLOWCHART FINALE>, dovrai costruire gli elementi basandoti esclusivamente sulle informazioni presenti nell'<ARTICOLO DI LEGGE> e nel <FLOWCHART> forniti.).

**Esempio di Task con una sola azione**:
    {
      "soggetto": "Responsabile del procedimento",
      "azioni": [
        "Predisposizione del verbale"
      ]
    }

**Esempio di Task con più azioni**:
    {
      "soggetto": "Ente pubblico",
      "azioni": [
        "Istruttoria preliminare",
        "Verifica completezza della documentazione"
      ]
    }


**Esempio di Gateway e ConditionalFlow**:
    {
      "domanda": "Documentazione completa?",
      "rami": [
        {
          "risposta": "sì",
          "ramo": [
            {
              "soggetto": "Ente pubblico",
              "azioni": [
                "Analisi della documentazione"
              ]
            }
          ],
          "esito": "CONTINUA"
        },
        {
          "risposta": "no",
          "ramo": [],
          "esito": "FINE"
        }
      ]
    }


**Esempio di Gateway e ConditionalFlow con Gateways annidati**:
    {
      "domanda": "Parere favorevole della Conferenza dei Servizi?",
      "rami": [
        {
          "risposta": "sì",
          "ramo": [
            {
              "soggetto": "Ufficio tecnico",
              "azioni": [
                "Redazione del provvedimento finale",
                "Trasmissione del provvedimento al richiedente"
              ]
            }
          ],
          "esito": "CONTINUA"
        },
        {
          "risposta": "no",
          "ramo": [
            {
              "domanda": "Presentazione osservazioni correttive?",
              "rami": [
                {
                  "risposta": "sì",
                  "ramo": [
                    {
                      "soggetto": "Commissione tecnica",
                      "azioni": [
                        "Valutazione delle osservazioni",
                        "Revisione del progetto"
                      ]
                    }
                  ],
                  "esito": "CONTINUA"
                },
                {
                  "risposta": "no",
                  "ramo": [
                    {
                      "soggetto": "Ufficio procedimenti",
                      "azioni": [
                        "Comunicazione di diniego motivato",
                        "Archiviazione della pratica"
                      ]
                    }
                  ],
                  "esito": "FINE"
                }
              ]
            }
          ],
          "esito": "EREDITATO"
        }
      ]
    }
"""
BACK = """#SYSTEM
**Contesto**: Sei un esperto nell’analisi e controllo di procedure amministrative.

**Obiettivo**: Leggi attentamente un <ARTICOLO DI LEGGE> e verifica la correttezza di un flowchart che descrive la relativa procedura amministrativa. Se necessario, correggi gli errori e fornisci il flowchart definitivo, rappresentato in formato JSON preciso e conforme.

**Input**:
    - <ARTICOLO DI LEGGE>: il testo normativo che descrive la procedura amministrativa.
    - <FLOWCHART>: il flowchart testuale della procedura amministrativa, estratto dall'<ARTICOLO DI LEGGE>.

**Output**: 
    <FLOWCHART DEFINITIVO>: il flowchart corretto e definitivo, rappresentato in formato JSON conforme alla seguente struttura concettuale:

        - Task:
            - "soggetto" (stringa): Chi compie l'azione (es. "Ente pubblico").
            - "azioni" (lista di stringhe): Le azioni svolte in sequenza.

        - Gateway:
            - "domanda" (stringa): La domanda o il punto decisionale.
            - "rami" (lista di ConditionalFlow): Le alternative in base alla risposta.

        - ConditionalFlow:
            - "risposta" (stringa): La risposta o condizione.
            - "ramo" (lista ricorsiva): Gli elementi successivi (Task o ulteriori Condizioni).
            - "esito" (stringa): CONTINUA | FINE | EREDITATO

        - Flowchart:
            - Lista ordinata e strutturata degli elementi della procedura.

**Approccio STEP-BY-STEP**:
    Segui rigorosamente questi passaggi:
    1. Leggi attentamente l’<ARTICOLO DI LEGGE> per comprenderne il contesto e la sequenza logica della procedura;
    2. Individua le Task, Gateway e ConditionalFlow che descrivono la procedura amministrativa;
    2. Analizza il <FLOWCHART> proposto;
    3. Confronta in modo preciso e rigoroso il <FLOWCHART> con la procedura dall'<ARTICOLO DI LEGGE> che hai estratto;
    4. Per ogni elemento del <FLOWCHART>:
        a. Verifica se è:
            - Corretto: lo lasci invariato;
            - Errato, impreciso o incompleto: lo correggi;
            - Mancante: lo aggiungi;
            - Superfluo: lo elimini;
    4. Controllo finale:
        a. Verifica completezza, coerenza e conformità al formato richiesto;
        b. Assicurati che il <FLOWCHART FINALE> descriva perfettamente la procedura amministrativa;
        c. Non introdurre interpretazioni personali o elementi non esplicitamente presenti nell'<ARTICOLO DI LEGGE>.

**Vincoli e Stile**:
    - Formula azioni: Descrivi le azioni nella forma "nome deverbale (azione) + complemento di specificazione"
    - Descrizioni sintetiche: Utilizza descrizioni brevi ma chiare e precis
    - Formato preciso: Il <FLOWCHART FINALE> deve essere conforme alla struttura specificata

**Audience**: Il destinatario è un analista tecnico di procedure amministrative, quindi il <FLOWCHART FINALE> deve essere preciso, corretto e pronto per la modellazione grafica.

#ESEMPIO
Di seguito un esempio di flowchart strutturato in formato JSON, conforme allo schema previsto. L'esempio deriva da un procedimento amministrativo ipotetico e mostra la corretta struttura e annidamento degli oggetti.
**ESEMPIO DI <FLOWCHART STRUTTURATO>**:
[
  {
    "soggetto": "Ente pubblico",
    "azioni": [
      "Avvio dell'istruttoria preliminare",
      "Verifica completezza della documentazione"
    ]
  },
  {
    "domanda": "Documentazione completa?",
    "rami": [
      {
        "risposta": "sì",
        "ramo": [],
        "esito": "CONTINUA"
      },
      {
        "risposta": "no",
        "ramo": [
          {
            "soggetto": "Ente pubblico",
            "azioni": [
              "Richiesta integrazione della documentazione"
            ]
          },
          {
            "domanda": "Documentazione inviata in tempo?",
            "rami": [
              {
                "risposta": "sì",
                "ramo": [
                  {
                    "soggetto": "Ente pubblico",
                    "azioni": [
                      "Comunicazione di ricevuta integrazione"
                    ]
                  }
                ],
                "esito": "CONTINUA"
              },
              {
                "risposta": "no",
                "ramo": [
                  {
                    "soggetto": "Ente pubblico",
                    "azioni": [
                      "Archiviazione della richiesta"
                    ]
                  }
                ],
                "esito": "FINE"
              }
            ]
          }
        ],
        "esito": "EREDITATO"
      }
    ]
  },
  {
    "soggetto": "Ente pubblico",
    "azioni": [
      "Valutazione del progetto",
      "Redazione della relazione tecnica"
    ]
  },
  {
    "soggetto": "Consiglio direttivo",
    "azioni": [
      "Valutazione ammissibilità",
      "Adozione della delibera"
    ]
  }
]"""
human_msg_2 = """##INPUT
<ARTICOLO DI LEGGE>:
{law_article}
<FLOWCHART>:
{flowchart}
##OUTPUT
<FLOWCHART DEFINITIVO>:"""

prompt_2 = ChatPromptTemplate.from_messages([
    SystemMessagePromptTemplate.from_template(system_msg_2),
    HumanMessagePromptTemplate.from_template(human_msg_2)
])

chain2 = LLMChain(llm=llm, prompt=prompt_2, output_key="updated_flowchart")




class Metadati(BaseModel):
    # Campi forniti in input (obbligatori)
    Attività: str = Field(..., description="(Lo ricevi in INPUT) il nome dell'attività")
    ID: str = Field(..., description="(Lo ricevi in INPUT) identificatore dell'attività")
    attore: str = Field(..., description="(Lo ricevi in INPUT) chi esegue l'attività")

    
    # Campi da completare (obbligatori)
    descrizione: str = Field(..., description="Breve descrizione dell'attività")
    riferimento_normativo: str = Field(..., description="Riferimento normativo puntuale (riferimento normativo più preciso possibile)")
    destinatario: str = Field(..., description="Entità che riceve l'effetto dell'attività")
    output: str = Field(..., description="Risultato, richiesta o decisione finale dell’attività")
    
    # Campi opzionali
    documenti: Optional[str] = Field(None, description="Documenti coinvolti nell'attività")
    riferimento_temporale: Optional[str] = Field(None, description="Durata stimata per completare l'attività o limite massimo")
    moduli: Optional[str] = Field(None, description="Moduli utilizzati durante l'attività")

class ActivityCompletion(BaseModel):
    completions: list[Metadati]
        
ActivityCompletion.model_rebuild()

back = """
#ESEMPIO DI ELEMENTI DEL <JSON FINALE>:  
    Di seguito ti fornisco alcuni esempi di elementi del <JSON FINALE>, in formato JSON conforme allo schema previsto.  
    Gli esempi servono solo per mostrarti alcune strutture tecniche e la forma corretta dei campi, non il contenuto specifico (per il contenuto del <JSON FINALE>, dovrai completare gli elementi basandoti esclusivamente sulle informazioni presenti nell'<ARTICOLO DI LEGGE> e nel <JSON DELLE ATTIVITÀ> forniti).  

**Esempio di attività con tutti i metadati valorizzati**:  
    {
      "Attività": "Verificare completezza documentazione",
      "ID": "art5_comma2",
      "attore": "Ufficio Protocollo",
      "descrizione": "Controllo della completezza della documentazione ricevuta",
      "riferimento_normativo": "Art. 5, comma 2",
      "destinatario": "Richiedente",
      "output": "Conferma o richiesta di integrazione",
      "documenti": "N/A",
      "riferimento_temporale": "entro 30 giorni",
      "moduli": "Modulo A-23"
    }  
"""
system_msg_3 = """#SYSTEM
**Contesto**: Sei un esperto nella strutturazione e completamento di metadati relativi a procedure amministrative.

**Obiettivo**: Leggi attentamente un <ARTICOLO DI LEGGE> e completa i metadati richiesti per ciascuna attività amministrativa, aggiornando un JSON precompilato.

**Input**:
    - <ARTICOLO DI LEGGE>: Il testo normativo o regolamentare di riferimento;
    - <JSON DELLE ATTIVITÀ>: Elenco delle attività da completare. Ogni attività contiene:
        - Attività;
        - ID;
        - Attore.

**Output**:  
Restituisci il <JSON FINALE>, in formato JSON conforme alla seguente struttura concettuale e ottenuto completando <JSON DELLE ATTIVITÀ> con i seguenti metadati:

    - Metadati obbligatori:
        - descrizione;
        - riferimento_normativo;
        - destinatario;
        - output.

    - Metadati opzionali:
        - documenti;
        - riferimento_temporale;
        - moduli.

**Approccio STEP-BY-STEP**:
    Segui rigorosamente questi passaggi:
    1. Leggi attentamente l’<ARTICOLO DI LEGGE> nella sua interezza per comprendere i dettagli delle attività;
    2. Per ogni attività A contenuta in <JSON DELLE ATTIVITÀ>:
        a. Ricerca nell’<ARTICOLO DI LEGGE> i metadati richiesti;
        b. Per ogni metadato dell'Attività A:
            - Se il metadato è presente o chiaramente ricavabile:
                - Inserisci il valore corretto nel <JSON FINALE>;
            - Altrimenti (solo per i metadati opzionali):
                - Inserisci "N/A";
    3. Assicurati che i metadati dell'<JSON FINALE> siano completi, coerenti e conformi all’<ARTICOLO DI LEGGE>;

**Vincoli e Stile**:
    - Non lasciare campi obbligatori vuoti o con "N/A";
    - Solo i metadati opzionali possono essere valorizzati con "N/A" in assenza di informazioni;
    - Massima precisione nell'attribuzione dei riferimenti normativi;
    - Il <JSON FINALE> deve essere ordinato, chiaro e facilmente leggibile;
    - Non aggiungere interpretazioni personali o elementi non presenti nell'<ARTICOLO DI LEGGE>;

**Audience**: Il destinatario è uno specialista in gestione documentale e amministrativa, quindi il <JSON FINALE> deve essere impeccabile, strutturato e pronto per essere integrato nei sistemi documentali.
#ESEMPIO DI ELEMENTI DEL <JSON FINALE>:  
    Di seguito ti fornisco un esempio di Attività del <JSON FINALE>, in formato JSON conforme allo schema previsto.  
    L'esempio solo per mostrarti la struttura tecnica e la forma corretta dei campi, non il contenuto specifico (per il contenuto del <JSON FINALE>, dovrai completare gli elementi basandoti esclusivamente sulle informazioni presenti nell'<ARTICOLO DI LEGGE> e nel <JSON DELLE ATTIVITÀ> forniti).  

**Esempio di attività con tutti i metadati valorizzati**:  
    {
      "Attività": "Verificare completezza documentazione",
      "ID": "A9",
      "attore": "Ufficio Protocollo",
      "descrizione": "Controllo della completezza della documentazione ricevuta",
      "riferimento_normativo": "Art. 5, comma 2",
      "destinatario": "Richiedente",
      "output": "Conferma o richiesta di integrazione",
      "documenti": "N/A",
      "riferimento_temporale": "entro 30 giorni",
      "moduli": "Modulo A-23"
    }
"""


##RUN


path_base = '' #INSERT THE PATH CONTAINING THE PROCEDURES IN .WORD FORMAT
paths = glob(os.path.join(path_base, "*.docx"))

for percorso_file in paths:
    nome = os.path.splitext(os.path.basename(percorso_file))[0]
    try:
        doc = Document(percorso_file)
        law_article = '\n'.join([par.text for par in doc.paragraphs])
    except Exception as e:
        print(f"Errore nel leggere {percorso_file}: {e}")
    
    print(f'{nome}: {len(law_article)}')
    
    completion1 = client.chat.completions.create(
            model="gpt-4o-2024-08-06",
            messages=[
                {"role": "system", "content": system_msg_1},
                {"role": "user", "content": f"""
#INPUT:
<ARTICOLO DI LEGGE>:
    {law_article}

#OUTPUT:"""}
            ],
            temperature=0
        )

    flowchart_testuale = completion1.choices[0].message.content.strip()

    completion2 = client.beta.chat.completions.parse(
            model="gpt-4o-2024-08-06",
            messages=[
                {"role": "system", "content": system_msg_2},
                {"role": "user", "content": f"""
#INPUT:
<ARTICOLO DI LEGGE>:
    {law_article}
<FLOWCHART>:
    {flowchart_testuale}

#OUTPUT:
"""}
            ],
            response_format=Flowchart,
            temperature=0
        )

    output = completion2.choices[0].message.parsed
    flowchart = [step.model_dump() for step in output.passaggi]

    add_dummy(flowchart)
    add_ids(flowchart)
    add_successor_ids(flowchart)
    df = json_to_dataframe(flowchart)

    json_list = []
    filtered_df = df[~(df["ID"].str.startswith(("R", "C"))) &  (df["NOME"] != "Fine") &  (df["NOME"] != "Continua")][["NOME", "ID", "SOGGETTO"]]

    for _, row in filtered_df.iterrows():
        metadato = {
            "Attività": row["NOME"],
            "ID":row["ID"],
            "Attore": row["SOGGETTO"]
        }
        json_list.append(metadato)

    json_output = json.dumps(json_list, ensure_ascii=False, indent=2)
    json_list = []
    filtered_df = df[~(df["ID"].str.startswith(("R", "C"))) &  (df["NOME"] != "Fine") &  (df["NOME"] != "Continua")][["NOME", "ID", "SOGGETTO"]]

    for _, row in filtered_df.iterrows():
        metadato = {
            "Attività": row["NOME"],
            "ID":row["ID"],
            "Attore": row["SOGGETTO"]
        }
        json_list.append(metadato)

    json_output = json.dumps(json_list, ensure_ascii=False, indent=2)

    completion3 = client.beta.chat.completions.parse(
            model="gpt-4o-2024-08-06",
            messages=[
                {"role": "system", "content": system_msg_3},
                {"role": "user", "content": f"""
#INPUT:
<ARTICOLO DI LEGGE>:
    {law_article}
<JSON DELLE ATTIVITÀ>:
    {json_output}

#OUTPUT:
"""}
            ],
            response_format=ActivityCompletion,
            temperature=0
        )
    res3 = completion3.choices[0].message.parsed




    flowchart_metadati = [step.model_dump() for step in res3.completions]

    metadati_json = pd.DataFrame(flowchart_metadati)
    metadati_json = metadati_json.rename(columns={"ID": "ID_y"})

    df_merged = df.merge(metadati_json, left_on="ID", right_on="ID_y", how="left")

    df_merged = df_merged.drop(columns=["ID_y"])
    df_merged

    path =f"" ##INSERT OUTPUT PATH

    bpmn_content, html_content = create_bpmn_from_df(df_merged, f"{nome}.bpmn")  
    csv_finale = df_merged  
    bpmn_finale = bpmn_content  
    html_finale = html_content  


    salva_documenti(csv_finale, bpmn_finale, html_finale, nome, path)





